
from flask import render_template, request, redirect, session, send_file, jsonify
from database import connect
from datetime import datetime
import math
import os
from openpyxl import Workbook
from fyers_apiv3 import fyersModel

LOT_SIZES = {
    "NIFTY": 60,
    "BANKNIFTY": 30,
    "FINNIFTY": 65
}


# ================= FYERS CLIENT =================
def get_fyers_client():
    token = session.get("fyers_access_token")

    if not token:
        return None

    client_id = os.getenv("FYERS_APP_ID")

    return fyersModel.FyersModel(
        client_id=client_id,
        token=token,
        log_path=""
    )


# ================= LOT LOGIC =================
def get_lot(symbol):
    symbol = symbol.upper()
    for k in LOT_SIZES:
        if k in symbol:
            return LOT_SIZES[k]
    return None


def calc_qty(symbol, entry, capital):
    lot = get_lot(symbol)

    if lot:
        one = entry * lot
        lots = math.floor(capital / one)
        if lots < 1:
            lots = 1
        return lots * lot

    qty = math.floor(capital / entry)
    return max(qty, 1)


# ================= ROUTES =================
def trade_routes(app):

    # ================= DASHBOARD =================
    @app.route("/dashboard")
    def dashboard():

        if "user" not in session:
            return redirect("/")
        
        if "fyers_access_token" not in session:
            return redirect("fyers-login")

        conn = connect()
        cur = conn.cursor()

        cur.execute("""
        SELECT trade_mode, capital, parts,
        trail_on, trail_trigger,
        trail_move, breakeven_trigger,
        live_master
        FROM settings WHERE id=1
        """)

        s = cur.fetchone()

        trade_mode = s[0]
        capital = s[1]
        parts = s[2]
        trail_on = s[3]
        live_master = s[7]

        view_mode = "LIVE" if live_master == "ON" else "PAPER"

        cur.execute("""
        SELECT * FROM trades
        WHERE mode=?
        ORDER BY id ASC
        """, (view_mode,))

        rows = cur.fetchall()

        trades = []
        pnl_total = 0
        wins = 0
        losses = 0
        open_trades = 0
        equity = []
        running = 0

        for r in rows:
            pnl = r[10]
            status = r[9]

            pnl_total += pnl
            running += pnl
            equity.append(running)

            if pnl > 0:
                wins += 1
            elif pnl < 0:
                losses += 1

            if status == "OPEN":
                open_trades += 1

            trades.append({
                "id": r[0],
                "mode": r[1],
                "symbol": r[2],
                "side": r[3],
                "qty": r[4],
                "entry": r[5],
                "sl": r[6],
                "target": r[7],
                "status": r[9],
                "pnl": r[10]
            })

        total_closed = wins + losses
        win_rate = round((wins / total_closed) * 100, 2) if total_closed else 0

        conn.close()

        return render_template(
            "dashboard.html",
            trades=trades,
            total_trades=len(rows),
            wins=wins,
            losses=losses,
            win_rate=win_rate,
            open_trades=open_trades,
            total_pnl=round(pnl_total, 2),
            equity_curve=equity,
            capital=capital,
            parts=parts,
            trade_mode=trade_mode,
            trail_on=trail_on,
            live_master=live_master,
            view_mode=view_mode
        )

    # ================= FYERS SYNC =================

    @app.route("/sync-positions")
    def sync_positions():
        fyers = get_fyers_client()

        if not fyers:
            return jsonify({"error": "FYERS not logged in"})

        response = fyers.positions()
        return jsonify(response)

    @app.route("/sync-orders")
    def sync_orders():
        fyers = get_fyers_client()

        if not fyers:
            return jsonify({"error": "FYERS not logged in"})

        response = fyers.orderbook()
        return jsonify(response)

    @app.route("/get-ltp/<symbol>")
    def get_ltp(symbol):
        fyers = get_fyers_client()

        if not fyers:
            return jsonify({"error": "FYERS not logged in"})

        data = {"symbols": symbol}
        response = fyers.quotes(data)

        return jsonify(response)

    # ================= ADD TRADE =================
    @app.route("/add_trade", methods=["POST"])
    def add_trade():

        symbol = request.form["symbol"].upper()
        side = request.form["side"]
        entry = float(request.form["entry"])
        sl = float(request.form["sl"])
        rr = float(request.form["rr"])

        conn = connect()
        cur = conn.cursor()

        cur.execute("""
        SELECT capital, parts, live_master
        FROM settings WHERE id=1
        """)

        s = cur.fetchone()

        capital = s[0]
        parts = s[1]
        live_master = s[2]

        mode = "LIVE" if live_master == "ON" else "PAPER"

        if parts <= 0:
            parts = 1

        per_capital = capital / parts

        qty = calc_qty(symbol, entry, per_capital)

        if side == "LONG":
            target = entry + ((entry - sl) * rr)
        else:
            target = entry - ((sl - entry) * rr)

        now = datetime.now().strftime("%d-%m-%Y %H:%M")

        cur.execute("""
        INSERT INTO trades
        VALUES (NULL,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            mode,
            symbol,
            side,
            qty,
            entry,
            sl,
            round(target, 2),
            rr,
            "OPEN",
            0,
            now,
            ""
        ))

        conn.commit()
        conn.close()

        return redirect("/dashboard")

    # ================= RESET =================
    @app.route("/reset_trades")
    def reset_trades():

        conn = connect()
        cur = conn.cursor()

        cur.execute("SELECT live_master FROM settings WHERE id=1")
        live = cur.fetchone()[0]

        mode = "LIVE" if live == "ON" else "PAPER"

        cur.execute("DELETE FROM trades WHERE mode=?", (mode,))

        conn.commit()
        conn.close()

        return redirect("/dashboard")

    # ================= EXPORT =================
    @app.route("/export_csv")
    def export_csv():

        conn = connect()
        cur = conn.cursor()

        wb = Workbook()

        ws1 = wb.active
        ws1.title = "PAPER"

        ws2 = wb.create_sheet("LIVE")

        headers = [
            "ID","Mode","Symbol","Side","Qty",
            "Entry","SL","Target","RR","Status",
            "PnL","Entry Time","Exit Time"
        ]

        ws1.append(headers)
        ws2.append(headers)

        cur.execute("SELECT * FROM trades WHERE mode='PAPER'")
        for row in cur.fetchall():
            ws1.append(row)

        cur.execute("SELECT * FROM trades WHERE mode='LIVE'")
        for row in cur.fetchall():
            ws2.append(row)

        conn.close()

        file = "trades.xlsx"
        wb.save(file)

        return send_file(file, as_attachment=True)

from flask import render_template, request, redirect, session, send_file, jsonify
from database import connect
from datetime import datetime
import math
import os
from openpyxl import Workbook
from fyers_apiv3 import fyersModel

LOT_SIZES = {
    "NIFTY": 60,
    "BANKNIFTY": 30,
    "FINNIFTY": 65
}


# ================= FYERS CLIENT =================
def get_fyers_client():
    token = session.get("fyers_access_token")

    if not token:
        return None

    client_id = os.getenv("FYERS_APP_ID")

    return fyersModel.FyersModel(
        client_id=client_id,
        token=token,
        log_path=""
    )


# ================= LOT LOGIC =================
def get_lot(symbol):
    symbol = symbol.upper()
    for k in LOT_SIZES:
        if k in symbol:
            return LOT_SIZES[k]
    return None


def calc_qty(symbol, entry, capital):
    lot = get_lot(symbol)

    if lot:
        one = entry * lot
        lots = math.floor(capital / one)
        if lots < 1:
            lots = 1
        return lots * lot

    qty = math.floor(capital / entry)
    return max(qty, 1)


# ================= ROUTES =================
def trade_routes(app):

    # ================= DASHBOARD =================
    @app.route("/dashboard")
    def dashboard():

        if "user" not in session:
            return redirect("/")

        conn = connect()
        cur = conn.cursor()

        cur.execute("""
        SELECT trade_mode, capital, parts,
        trail_on, trail_trigger,
        trail_move, breakeven_trigger,
        live_master
        FROM settings WHERE id=1
        """)

        s = cur.fetchone()

        trade_mode = s[0]
        capital = s[1]
        parts = s[2]
        trail_on = s[3]
        live_master = s[7]

        view_mode = "LIVE" if live_master == "ON" else "PAPER"

        cur.execute("""
        SELECT * FROM trades
        WHERE mode=?
        ORDER BY id ASC
        """, (view_mode,))

        rows = cur.fetchall()

        trades = []
        pnl_total = 0
        wins = 0
        losses = 0
        open_trades = 0
        equity = []
        running = 0

        for r in rows:
            pnl = r[10]
            status = r[9]

            pnl_total += pnl
            running += pnl
            equity.append(running)

            if pnl > 0:
                wins += 1
            elif pnl < 0:
                losses += 1

            if status == "OPEN":
                open_trades += 1

            trades.append({
                "id": r[0],
                "mode": r[1],
                "symbol": r[2],
                "side": r[3],
                "qty": r[4],
                "entry": r[5],
                "sl": r[6],
                "target": r[7],
                "status": r[9],
                "pnl": r[10]
            })

        total_closed = wins + losses
        win_rate = round((wins / total_closed) * 100, 2) if total_closed else 0

        conn.close()

        return render_template(
            "dashboard.html",
            trades=trades,
            total_trades=len(rows),
            wins=wins,
            losses=losses,
            win_rate=win_rate,
            open_trades=open_trades,
            total_pnl=round(pnl_total, 2),
            equity_curve=equity,
            capital=capital,
            parts=parts,
            trade_mode=trade_mode,
            trail_on=trail_on,
            live_master=live_master,
            view_mode=view_mode
        )

    # ================= FYERS SYNC =================

    @app.route("/sync-positions")
    def sync_positions():
        fyers = get_fyers_client()

        if not fyers:
            return jsonify({"error": "FYERS not logged in"})

        response = fyers.positions()
        return jsonify(response)

    @app.route("/sync-orders")
    def sync_orders():
        fyers = get_fyers_client()

        if not fyers:
            return jsonify({"error": "FYERS not logged in"})

        response = fyers.orderbook()
        return jsonify(response)

    @app.route("/get-ltp/<symbol>")
    def get_ltp(symbol):
        fyers = get_fyers_client()

        if not fyers:
            return jsonify({"error": "FYERS not logged in"})

        data = {"symbols": symbol}
        response = fyers.quotes(data)

        return jsonify(response)

    # ================= ADD TRADE =================
    @app.route("/add_trade", methods=["POST"])
    def add_trade():

        symbol = request.form["symbol"].upper()
        side = request.form["side"]
        entry = float(request.form["entry"])
        sl = float(request.form["sl"])
        rr = float(request.form["rr"])

        conn = connect()
        cur = conn.cursor()

        cur.execute("""
        SELECT capital, parts, live_master
        FROM settings WHERE id=1
        """)

        s = cur.fetchone()

        capital = s[0]
        parts = s[1]
        live_master = s[2]

        mode = "LIVE" if live_master == "ON" else "PAPER"

        if parts <= 0:
            parts = 1

        per_capital = capital / parts

        qty = calc_qty(symbol, entry, per_capital)

        if side == "LONG":
            target = entry + ((entry - sl) * rr)
        else:
            target = entry - ((sl - entry) * rr)

        now = datetime.now().strftime("%d-%m-%Y %H:%M")

        cur.execute("""
        INSERT INTO trades
        VALUES (NULL,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            mode,
            symbol,
            side,
            qty,
            entry,
            sl,
            round(target, 2),
            rr,
            "OPEN",
            0,
            now,
            ""
        ))

        conn.commit()
        conn.close()

        return redirect("/dashboard")

    # ================= RESET =================
    @app.route("/reset_trades")
    def reset_trades():

        conn = connect()
        cur = conn.cursor()

        cur.execute("SELECT live_master FROM settings WHERE id=1")
        live = cur.fetchone()[0]

        mode = "LIVE" if live == "ON" else "PAPER"

        cur.execute("DELETE FROM trades WHERE mode=?", (mode,))

        conn.commit()
        conn.close()

        return redirect("/dashboard")

    # ================= EXPORT =================
    @app.route("/export_csv")
    def export_csv():

        conn = connect()
        cur = conn.cursor()

        wb = Workbook()

        ws1 = wb.active
        ws1.title = "PAPER"

        ws2 = wb.create_sheet("LIVE")

        headers = [
            "ID","Mode","Symbol","Side","Qty",
            "Entry","SL","Target","RR","Status",
            "PnL","Entry Time","Exit Time"
        ]

        ws1.append(headers)
        ws2.append(headers)

        cur.execute("SELECT * FROM trades WHERE mode='PAPER'")
        for row in cur.fetchall():
            ws1.append(row)

        cur.execute("SELECT * FROM trades WHERE mode='LIVE'")
        for row in cur.fetchall():
            ws2.append(row)

        conn.close()

        file = "trades.xlsx"
        wb.save(file)

        return send_file(file, as_attachment=True)

